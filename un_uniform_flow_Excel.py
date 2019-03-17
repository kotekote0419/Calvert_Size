import openpyxl as px


def xl_sheet(flow_data, flow_detail, case, d_t):
    wb = px.Workbook()
    ws = wb.create_sheet(title='CASE-{0}'.format(case+1))
    ws.cell(1, 1).value = 'Un-uniform flow Calculation: CASE-{0}'.format(case+1)
    ws.cell(3, 2).value = 'Q(m3/min)'
    ws.cell(3, 3).value = 'L(m)'
    ws.cell(3, 4).value = 'Down WL(m)'
    ws.cell(3, 5).value = 'Up WL(m)'
    for i in range(len(flow_data)):
        for j in range(5):
            ws.cell(i+4, j+1).value = flow_data[i][j]
    #for i in range(len(flow_detail)):
    #    for j in range(len(flow_detail[i])):
    #        ws.cell(j+1, i+6).value = flow_detail[i][j]
    img = px.drawing.image.Image('Calvert{0}.png'.format(case+1))
    ws.add_image(img, 'A{0}'.format(len(flow_data)+1))
    wb.save('xlSheet'+d_t+'.xls')
    print(ws.page_margins.left)
    return wb


def xl_update(wb, flow_data, case, d_t):
    ws = wb.create_sheet(title='CASE-{0}'.format(case + 1))
    ws.cell(1, 2).value = 'Q(m3/min)'
    ws.cell(1, 3).value = 'L(m)'
    ws.cell(1, 4).value = 'Down WL(m)'
    ws.cell(1, 5).value = 'Up WL(m)'
    for i in range(len(flow_data)):
        for j in range(5):
            ws.cell(i + 2, j + 1).value = flow_data[i][j]
    img = px.drawing.image.Image('Calvert{0}.png'.format(case + 1))
    ws.add_image(img, 'A{0}'.format(len(flow_data) + 1))
    wb.save('xlSheet' + d_t + '.xls')
    return wb
